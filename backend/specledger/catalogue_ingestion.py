"""Safe ingestion primitives for messy industrial catalogue files.

This module deliberately stops before enrichment. It turns a source table into
stable, traceable rows so later matching and generation can be deterministic.
It accepts CSV/TSV, JSON and XML, and optionally XLSX when ``openpyxl`` is
installed.
"""

from __future__ import annotations

import csv
import io
import hashlib
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Mapping
from xml.etree import ElementTree


PLACEHOLDERS = frozenset({"", "-", "--", "n/a", "na", "none", "null", "unknown"})


def canonical_key(value: str) -> str:
    """Create a stable comparison key without destroying the source value."""
    return re.sub(r"[^a-z0-9]+", "_", value.strip().casefold()).strip("_")


def clean_cell(value: object) -> str | None:
    """Normalize whitespace and remove known placeholder values."""
    if value is None:
        return None
    cleaned = re.sub(r"\s+", " ", str(value).replace("\u00a0", " ")).strip()
    return None if cleaned.casefold() in PLACEHOLDERS else cleaned


def clean_manufacturer_name(raw_name: str | None) -> str | None:
    """Strip distributor codes in parentheses, e.g. 'Freud Inc (2435)' -> 'Freud Inc'."""
    if not raw_name:
        return None
    # Remove code in trailing parentheses like (2435) or (JAMIN)
    cleaned = re.sub(r"\s*\([A-Z0-9_-]+\)\s*$", "", raw_name.strip(), flags=re.IGNORECASE).strip()
    return cleaned if cleaned else raw_name



@dataclass(frozen=True)
class SourceRow:
    row_number: int
    source_name: str
    source_fingerprint: str
    values: dict[str, str | None]


@dataclass(frozen=True)
class CatalogueBatch:
    source_name: str
    columns: tuple[str, ...]
    rows: tuple[SourceRow, ...]

    @property
    def row_count(self) -> int:
        return len(self.rows)


def row_fingerprint(values: Mapping[str, str | None]) -> str:
    payload = json.dumps(dict(sorted(values.items())), ensure_ascii=False, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def normalize_rows(source_name: str, rows: Iterable[Mapping[object, object]]) -> CatalogueBatch:
    """Normalize headers/cells while retaining row-level lineage."""
    materialized = list(rows)
    if not materialized:
        raise ValueError("Catalogue file contains no data rows")
    columns: list[str] = []
    for row in materialized:
        for raw_key in row:
            # csv.DictReader files any fields beyond the header under a None
            # key, as a list. Those have no column name, so they are not a
            # column — treating them as one produced a phantom "none" column
            # holding a stringified Python list, which would then be carried
            # into the delivered 252-column export.
            if raw_key is None:
                continue
            key = canonical_key(str(raw_key))
            if key and key not in columns:
                columns.append(key)
    if not columns:
        raise ValueError("Catalogue file contains no usable columns")
    normalized: list[SourceRow] = []
    for number, row in enumerate(materialized, start=2):
        values = {
            canonical_key(str(key)): clean_cell(value)
            for key, value in row.items()
            if key is not None and canonical_key(str(key))
        }
        values = {column: values.get(column) for column in columns}
        # Skip rows that are entirely empty. Spreadsheets routinely carry
        # trailing blank rows, and a CSV can end with a stray newline; each
        # would otherwise become a product with no part number and no
        # description — a phantom SKU in the delivered catalogue.
        #
        # The counter still advances, so row numbers keep pointing at the
        # line they came from in the uploaded file. Lineage is the point.
        if not any(value for value in values.values()):
            continue
        normalized.append(SourceRow(number, source_name, row_fingerprint(values), values))
    if not normalized:
        raise ValueError("Catalogue file contains no data rows")
    return CatalogueBatch(source_name, tuple(columns), tuple(normalized))


# Encodings tried in order when reading a delimited file. UTF-8 covers most
# exports; cp1252 is what Excel on Windows writes by default, which is a very
# common way for a real catalogue to arrive. Latin-1 accepts any byte
# sequence, so it is last and only prevents a hard failure.
_TEXT_ENCODINGS = ("utf-8-sig", "cp1252", "latin-1")


def _decode_text(source: Path) -> str:
    """Read a delimited file, tolerating non-UTF-8 exports.

    Decoding as UTF-8 only meant a spreadsheet saved from Excel on Windows
    failed with a raw codec error naming a byte offset — accurate, and
    useless to whoever uploaded it.
    """
    raw = source.read_bytes()
    for encoding in _TEXT_ENCODINGS:
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError(
        "Could not decode this file as text. Save it as UTF-8 CSV "
        "(or upload the .xlsx directly) and try again."
    )


def _sniff_delimiter(text: str, default: str = ",") -> str:
    """Work out what separates the columns in a .csv file.

    Excel writes ";"-separated files in any locale where the comma is the
    decimal separator, and assuming a comma did not fail loudly: the file
    ingested, reported its row count, and put the whole line into a single
    column named after the joined header, so the delivered record was
    nonsense while the upload looked like it had worked.

    Decided on the header line alone, and only among characters that
    actually appear there, so a description containing commas cannot drag
    the choice away from the real separator. Ties go to the default.
    """
    header = next((line for line in text.splitlines() if line.strip()), "")
    if not header:
        return default
    candidates = (default, ";", "\t", "|")
    # csv.reader respects quoting, so a delimiter inside a quoted value is
    # not counted — which is the case a bare str.count() gets wrong.
    best, best_columns = default, 0
    for candidate in candidates:
        try:
            columns = len(next(csv.reader([header], delimiter=candidate)))
        except csv.Error:
            continue
        if columns > best_columns:
            best, best_columns = candidate, columns
    return best


def read_catalogue(path: str | Path, sheet_name: str | None = None) -> CatalogueBatch:
    """Read CSV, TSV, or XLSX into the normalized batch contract."""
    source = Path(path)
    suffix = source.suffix.casefold()
    if suffix in {".csv", ".tsv"}:
        text = _decode_text(source)
        delimiter = "\t" if suffix == ".tsv" else _sniff_delimiter(text)
        return normalize_rows(
            source.name, csv.DictReader(io.StringIO(text, newline=""), delimiter=delimiter)
        )
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("XLSX ingestion requires the openpyxl dependency") from exc
        workbook = load_workbook(source, read_only=True, data_only=True)
        worksheet = workbook[sheet_name] if sheet_name else workbook.active
        records = worksheet.iter_rows(values_only=True)
        headers = next(records, None)
        if not headers:
            raise ValueError("XLSX sheet contains no header row")
        return normalize_rows(source.name, (dict(zip(headers, values)) for values in records))
    if suffix == ".json":
        return normalize_rows(source.name, _read_json_records(_decode_text(source)))
    if suffix == ".xml":
        return normalize_rows(source.name, _read_xml_records(_decode_text(source)))
    raise ValueError(
        "Unsupported catalogue format; use CSV, TSV, XLSX, JSON, or XML"
    )


# A feed's product list is not always at the top level, and the wrapper key
# is whatever the publisher chose. These are the ones seen in the wild;
# anything else falls back to the first list of objects in the payload.
_JSON_LIST_KEYS = ("products", "items", "rows", "records", "data", "catalogue", "catalog")


def _read_json_records(text: str) -> list[dict]:
    """Pull a list of product objects out of a JSON feed."""
    try:
        payload = json.loads(text)
    except json.JSONDecodeError as exc:
        raise ValueError(f"File is not valid JSON: {exc}") from exc

    records = None
    if isinstance(payload, list):
        records = payload
    elif isinstance(payload, dict):
        for key in _JSON_LIST_KEYS:
            value = payload.get(key)
            if isinstance(value, list):
                records = value
                break
        if records is None:
            # A wrapper under an unrecognised key, or a single product.
            nested = [v for v in payload.values() if isinstance(v, list) and v]
            records = nested[0] if nested else [payload]

    if not isinstance(records, list) or not records:
        raise ValueError("JSON feed contains no product records")
    if not all(isinstance(record, dict) for record in records):
        raise ValueError("JSON feed must be a list of objects, one per product")
    return [_flatten_record(record) for record in records]


def _flatten_record(record: dict) -> dict:
    """Render nested values as text rather than dropping the field.

    A cell holds one value. A nested object or list still came from the
    supplier, so it is preserved as its JSON text and left for a human
    rather than silently discarded.
    """
    flat: dict = {}
    for key, value in record.items():
        if isinstance(value, (dict, list)):
            flat[key] = json.dumps(value, ensure_ascii=False)
        else:
            flat[key] = value
    return flat


def _read_xml_records(text: str) -> list[dict]:
    """Pull a list of product elements out of an XML feed.

    Entity declarations are refused before parsing rather than trusted:
    ElementTree's parser expands internal entities, which is how a small
    file becomes an out-of-memory error.
    """
    lowered = text.casefold()
    if "<!entity" in lowered or "<!doctype" in lowered:
        raise ValueError("XML with a DTD or entity declarations is not accepted")

    try:
        root = ElementTree.fromstring(text)
    except ElementTree.ParseError as exc:
        raise ValueError(f"File is not valid XML: {exc}") from exc

    records: list[dict] = []
    for element in list(root):
        record: dict = dict(element.attrib)
        for child in element:
            tag = child.tag.rsplit("}", 1)[-1]  # drop any namespace
            record[tag] = (child.text or "").strip()
        if element.text and element.text.strip() and not record:
            record[root.tag.rsplit("}", 1)[-1]] = element.text.strip()
        if record:
            records.append(record)

    if not records:
        raise ValueError("XML feed contains no product elements")
    return records
